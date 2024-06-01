import openpyxl
import singo


# 필요한 초기 정보 입력
file_path = input("===== 대상자 파일 경로 입력:") + ".xlsx"
start_num = int(input("===== 시작 행 번호 입력: ").replace(" ", ""))
end_num = int(input("===== 마지막 행 번호 입력: ").replace(" ", ""))


# 엑셀 항목의 첫 번째 시트에 접근
try:
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
except Exception as e:
    print("===== ", e, " =====")
    print("===== 파일 제목과 시트를 확인해주세요 =====\n")
    exit()


# 카테고리는 1번째 row에 존재
cate_tp = ws[1]


# 각 카테고리의 위치 저장
cate_idx_dic = {}
for idx in range(0, len(cate_tp)):
    cate_idx_dic[cate_tp[idx].value.replace(" ", "")] = idx


row_num = start_num
while (row_num <= end_num):
    target_info_tp = ws[row_num]

    target_info_dic = {
         
         "HT_TIN" :         target_info_tp[cate_idx_dic.get("HT_TIN")].value,
         "USER_NM":         target_info_tp[cate_idx_dic.get("이름")].value,
         "HP_NO":           target_info_tp[cate_idx_dic.get("연락처")].value,
         "CORP_NM":         target_info_tp[cate_idx_dic.get("사업장명")].value,
         "CORP_NO":         target_info_tp[cate_idx_dic.get("사업자번호")].value,
         "CORP_ADDR":       target_info_tp[cate_idx_dic.get("도로명주소")].value,
         "CORP_LB_ADDR":    target_info_tp[cate_idx_dic.get("지번주소")].value,
         "HTX_REG_NO":      target_info_tp[cate_idx_dic.get("홈택스접수번호")].value,
         "TOT_HTAX_AMT":    target_info_tp[cate_idx_dic.get("원천세신고금액")].value,
         "UUID":            target_info_tp[0].value.replace(" ", "")
    }

    target = singo.CorpSingo(target_info_dic)
    print("\n=========  row_num :: ", row_num, " ============\n")
    print("HT_TIN: "            , target.ht_tin)
    print("이름: "              , target.user_nm)
    print("연락처: "            , target.hp_no)
    print("사업장명: "          , target.corp_nm)
    print("사업자번호: "        , target.corp_no)
    print("도로명주소: "        , target.corp_addr)
    print("도로명주소: "        , target.corp_lb_addr)
    print("홈택스접수번호: "    , target.htx_reg_no)
    print("원천세신고금액: "    , target.tot_htax_amt)
    print("\n=====================\n")
    print("https://sadr.ook.kr/user_detail.html?user=" + target.ht_tin)

    target.singo()

    row_num += 1